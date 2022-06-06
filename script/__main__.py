# *******************************************************************************************
#  File:  __main__.py
#
#  Created: 06-06-2022
#
#  Copyright (c) 2022 James Dooley <james@dooley.ch>
#
#  History:
#  06-06-2022: Initial version
#
# *******************************************************************************************

__author__ = "James Dooley"
__license__ = "MIT"
__version__ = "1.0.0"
__maintainer__ = "James Dooley"
__status__ = "Production"

import csv
from pathlib import Path
import openpyxl
import json
from attrs import frozen, field, asdict
from attrs.validators import instance_of, gt


@frozen
class Item:
    id: int = field(validator=[instance_of(int), gt(0)])
    name: str = field(validator=[instance_of(str)])


@frozen
class SubIndustry:
    id: int = field(validator=[instance_of(int), gt(0)])
    name: str = field(validator=[instance_of(str)])
    items: list[Item] = field(repr=False, factory=list)


@frozen
class Industry:
    id: int = field(validator=[instance_of(int), gt(0)])
    name: str = field(validator=[instance_of(str)])
    items: list[SubIndustry] = field(repr=False, factory=list)


@frozen
class IndustryGroup:
    id: int = field(validator=[instance_of(int), gt(0)])
    name: str = field(validator=[instance_of(str)])
    items: list[Industry] = field(repr=False, factory=list)


@frozen
class Sector:
    id: int = field(validator=[instance_of(int), gt(0)])
    name: str = field(validator=[instance_of(str)])
    items: list[IndustryGroup] = field(repr=False, factory=list)


class SectorEncoder(json.JSONEncoder):
    def default(self, value):
        return asdict(value)

def load_sub_industries() -> dict[int, list[Item]]:
    csv_file = Path('../csv/sub_industries.csv')

    sub_industries: dict[int, list[Item]] = dict()

    with csv_file.open('r') as file:
        reader = csv.reader(file)
        next(reader)
        for row in reader:
            id = int(row[0])
            industry_id = int(row[1])
            name = row[2]

            item = Item(id, name)

            if industry_id not in sub_industries:
                sub_industries[industry_id] = list()

            sub_industries[industry_id].append(item)
    return sub_industries


def load_industries(sub_industries: dict[int, list[Item]]) -> dict[int, list[Industry]]:
    csv_file = Path('../csv/industries.csv')

    industries: dict[int, list[Industry]] = dict()

    with csv_file.open('r') as file:
        reader = csv.reader(file)
        next(reader)
        for row in reader:
            id = int(row[0])
            group_industry_id = int(row[1])
            name = row[2]

            if group_industry_id not in industries:
                industries[group_industry_id] = list()

            industry = Industry(id, name)
            if id in sub_industries:
                subs = sub_industries[id]
                industry.items.extend(subs)

            industries[group_industry_id].append(industry)
    return industries


def load_industry_groups(industries: dict[int, list[Industry]]) -> dict[int, list[IndustryGroup]]:
    csv_file = Path('../csv/industry_groups.csv')

    groups: dict[int, list[IndustryGroup]] = dict()

    with csv_file.open('r') as file:
        reader = csv.reader(file)
        next(reader)
        for row in reader:
            id = int(row[0])
            sector_id = int(row[1])
            name = row[2]

            if sector_id not in groups:
                groups[sector_id] = list()

            group = IndustryGroup(id, name)
            if id in industries:
                inds = industries[id]
                group.items.extend(inds)

            groups[sector_id].append(group)

    return groups


def load_sectors(industry_groups: dict[int, list[IndustryGroup]]) -> list[Sector]:
    csv_file = Path('../csv/sectors.csv')

    sectors: list[Sector] = list()

    with csv_file.open('r') as file:
        reader = csv.reader(file)
        next(reader)
        for row in reader:
            id = int(row[0])
            name = row[1]

            sector = Sector(id, name)
            if id in industry_groups:
                groups = industry_groups[id]
                sector.items.extend(groups)
            sectors.append(sector)
    return sectors


def write_json_file() -> None:
    csv_file = Path('../gics.json')
    if csv_file.exists():
        csv_file.unlink()

    sub_industries = load_sub_industries()
    industries = load_industries(sub_industries)
    industry_groups = load_industry_groups(industries)
    sectors = load_sectors(industry_groups)

    json_content = json.dumps(sectors, indent=4, cls=SectorEncoder)

    csv_file.write_text(json_content)



def write_sectors_file() -> None:
    csv_file = Path('../csv/sectors.csv')
    if csv_file.exists():
        csv_file.unlink()

    excel_file = Path('../gics.xlsx')
    workbook: openpyxl.Workbook = openpyxl.load_workbook(excel_file)

    sheet = workbook['Sector']
    cells = sheet['A2':'B12']

    with csv_file.open('w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(['Id', 'Name'])
        for c1, c2 in cells:
            writer.writerow([c1.value, c2.value])


def write_industry_groups_file() -> None:
    csv_file = Path('../csv/industry_groups.csv')
    if csv_file.exists():
        csv_file.unlink()

    excel_file = Path('../gics.xlsx')
    workbook: openpyxl.Workbook = openpyxl.load_workbook(excel_file)

    sheet = workbook['Industry Group']
    cells = sheet['A2':'C25']

    with csv_file.open('w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(['Id', 'Sector Id', 'Name'])
        for c1, c2, c3 in cells:
            writer.writerow([c2.value, c1.value, c3.value])


def write_industries_file() -> None:
    csv_file = Path('../csv/industries.csv')
    if csv_file.exists():
        csv_file.unlink()

    excel_file = Path('../gics.xlsx')
    workbook: openpyxl.Workbook = openpyxl.load_workbook(excel_file)

    sheet = workbook['Industry']
    cells = sheet['A2':'C70']

    with csv_file.open('w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(['Id', 'Industry Group Id', 'Name'])
        for c1, c2, c3 in cells:
            writer.writerow([c2.value, c1.value, c3.value])


def write_sub_industries_file() -> None:
    csv_file = Path('../csv/sub_industries.csv')
    if csv_file.exists():
        csv_file.unlink()

    excel_file = Path('../gics.xlsx')
    workbook: openpyxl.Workbook = openpyxl.load_workbook(excel_file)

    sheet = workbook['Sub Industry']
    cells = sheet['A2':'C157']

    with csv_file.open('w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(['Id', 'Industry Id', 'Name'])
        for c1, c2, c3 in cells:
            writer.writerow([c2.value, c1.value, c3.value])


def main():
    # Create csv files
    write_sectors_file()
    write_industry_groups_file()
    write_industries_file()
    write_sub_industries_file()

    # JSON version
    write_json_file()

if __name__ == '__main__':
    main()
